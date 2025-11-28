"""
데이터 처리 및 Excel 내보내기 모듈 (완성본)
크롤링된 매출/매입 데이터를 분석하고 Excel 보고서를 생성합니다.
- 상세 내역을 '매출내역'과 '매입내역' 시트로 분리
- 월별 요약/손익 분석 데이터를 생성하여 별도 시트에 추가
"""

import os
from datetime import datetime
import pandas as pd
import numpy as np
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.utils import get_column_letter
import logging

logger = logging.getLogger(__name__)

# 상세 내역 시트 컬럼 순서
FINAL_DETAIL_COLUMNS = [
    '기안일', '문서제목', '기안부서', '문서번호', '링크', '구분',
    '거래처명', '공급가액', '부가세', '합계금액', '종결|완료'
]

def _prepare_detail_df(df: pd.DataFrame) -> pd.DataFrame:
    """
    상세 데이터를 Excel 내보내기에 적합한 형태로 정렬/가공
    """
    if df.empty:
        return pd.DataFrame(columns=FINAL_DETAIL_COLUMNS)

    prepared_df = df.copy()

    # 날짜 → 기안일
    if '날짜' in prepared_df.columns:
        prepared_df = prepared_df.rename(columns={'날짜': '기안일'})
    else:
        logger.error("❌ '날짜' 컬럼이 없어 처리 불가.")
        return pd.DataFrame(columns=FINAL_DETAIL_COLUMNS)

    # 금액 컬럼 정수 변환
    for col in ['공급가액', '부가세', '합계금액']:
        if col in prepared_df.columns:
            prepared_df[col] = pd.to_numeric(prepared_df[col], errors='coerce').fillna(0).astype(int)

    # 기안일 기준 정렬
    if '기안일' in prepared_df.columns:
        prepared_df = prepared_df.sort_values('기안일').reset_index(drop=True)

    # 최종 컬럼 순서 적용
    final_columns = [c for c in FINAL_DETAIL_COLUMNS if c in prepared_df.columns]
    return prepared_df[final_columns]

def process_monthly_summary(df: pd.DataFrame) -> pd.DataFrame:
    if df.empty:
        return pd.DataFrame(columns=['년월', '매출액', '매입액', '손익'])
    
    df_copy = df.copy()
    df_copy['년월'] = df_copy['날짜'].dt.to_period('M').astype(str)

    summary_df = df_copy.groupby(['년월', '구분'])['공급가액'].sum().unstack(fill_value=0).reset_index()
    summary_df = summary_df.rename(columns={'매출': '매출액', '매입': '매입액'})

    for col in ['년월', '매출액', '매입액']:
        if col not in summary_df.columns:
            summary_df[col] = 0

    summary_df['손익'] = summary_df['매출액'] - summary_df['매입액']
    summary_df = summary_df.sort_values('년월').reset_index(drop=True)
    
    return summary_df[['년월', '매출액', '매입액', '손익']]

def create_profit_analysis(monthly_df: pd.DataFrame) -> pd.DataFrame:
    if monthly_df.empty:
        return pd.DataFrame(columns=['년월','매출액','매입액','손익','누적손익','수익률','매출증감률','손익증감률'])
    
    df = monthly_df.copy()
    df['누적손익'] = df['손익'].cumsum()
    df['수익률'] = np.where(df['매출액'] > 0, (df['손익'] / df['매출액'] * 100).round(2), 0)
    df['매출증감률'] = df['매출액'].pct_change().fillna(0) * 100
    df['손익증감률'] = df['손익'].pct_change().fillna(0) * 100

    return df[['년월','매출액','매입액','손익','누적손익','수익률','매출증감률','손익증감률']]

def export_to_excel(detailed_df: pd.DataFrame, monthly_df: pd.DataFrame, analysis_df: pd.DataFrame, filename: str = None) -> str:
    """
    💡 핵심 변경점:
    ✅ 여기서 문서번호 → 기안부서 컬럼 생성
    """
    try:
        logger.info("📊 Excel 보고서 생성 시작")

        # ✅ 기안부서 생성 로직
        if '문서번호' in detailed_df.columns:
            detailed_df['기안부서'] = detailed_df['문서번호'].str.split('-', n=1).str[0].str.strip()

        output_dir = "output"
        if not os.path.exists(output_dir):
            os.makedirs(output_dir)

        if not filename:
            filename = f"매출매입현황_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        filepath = os.path.join(output_dir, filename)

        wb = Workbook()
        wb.remove(wb.active)

        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="366092", fill_type="solid")
        header_align = Alignment(horizontal="center", vertical="center")
        number_format = "#,##0"

        prepared_df = _prepare_detail_df(detailed_df)

        # 기간 상세
        if not prepared_df.empty:
            ws = wb.create_sheet("기간 상세 내역")
            add_dataframe_to_sheet(ws, prepared_df, "전체 거래 내역")
            format_worksheet(ws, prepared_df, header_font, header_fill, header_align, number_format)

        # 매출내역
        sales_df = prepared_df[(prepared_df['구분'] == '매출') & (prepared_df['종결|완료'].str.contains('종결', na=False))]
        if not sales_df.empty:
            ws = wb.create_sheet("매출내역")
            add_dataframe_to_sheet(ws, sales_df, "상세 매출 내역")
            format_worksheet(ws, sales_df, header_font, header_fill, header_align, number_format)

        # 매입내역
        purchase_df = prepared_df[(prepared_df['구분'] == '매입') & (prepared_df['종결|완료'].str.contains('종결', na=False))]
        if not purchase_df.empty:
            ws = wb.create_sheet("매입내역")
            add_dataframe_to_sheet(ws, purchase_df, "상세 매입 내역")
            format_worksheet(ws, purchase_df, header_font, header_fill, header_align, number_format)

        # 월별 요약
        if not monthly_df.empty:
            ws = wb.create_sheet("월별요약")
            add_dataframe_to_sheet(ws, monthly_df, "월별 매출/매입 요약")
            format_worksheet(ws, monthly_df, header_font, header_fill, header_align, number_format)

        # 손익 분석
        if not analysis_df.empty:
            ws = wb.create_sheet("손익분석")
            add_dataframe_to_sheet(ws, analysis_df, "손익 분석")
            format_worksheet(ws, analysis_df, header_font, header_fill, header_align, number_format)

        wb.save(filepath)
        logger.info(f"✅ Excel 생성 완료 → {filepath}")
        return filepath

    except Exception as e:
        logger.error(f"❌ Excel 생성 실패: {e}")
        raise

def add_dataframe_to_sheet(ws, df, title: str):
    ws['A1'] = title
    ws['A1'].font = Font(size=14, bold=True)

    for r in dataframe_to_rows(df, index=False, header=True):
        ws.append(r)

def format_worksheet(ws, df, header_font, header_fill, header_align, number_format):
    if df.empty: return

    for col in range(1, len(df.columns) + 1):
        cell = ws.cell(row=2, column=col)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align

    for col_idx, col_name in enumerate(df.columns, 1):
        is_amount = any(x in col_name for x in ['액','금액'])
        is_percent = any(x in col_name for x in ['률'])
        for row in range(3, len(df) + 3):
            cell = ws.cell(row=row, column=col_idx)
            if is_percent:
                cell.number_format = '0.00%'
            elif is_amount:
                cell.number_format = number_format

    for col_idx, column in enumerate(ws.columns, 1):
        max_length = 0
        column_letter = get_column_letter(col_idx)
        for cell in column:
            try:
                max_length = max(max_length, len(str(cell.value)))
            except:
                pass
        ws.column_dimensions[column_letter].width = min(max_length * 1.2 + 2, 50)
